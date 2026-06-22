"""
utils/report_generator.py
Genera reportes de resultados en PDF, XLSX y CSV.
"""

from __future__ import annotations
import csv
from datetime import datetime
from pathlib import Path
from typing import Sequence

from .logger import get_logger
from .file_utils import format_size
from models.history_entry import HistoryEntry

logger = get_logger("reports")


class ReportGenerator:
    """
    Genera reportes exportables de las comprensiones realizadas.

    Soporta: PDF (via reportlab), XLSX (via openpyxl), CSV (stdlib).
    """

    def export_csv(self, entries: Sequence[HistoryEntry], output_path: Path) -> bool:
        """
        Exporta entradas a CSV.

        Args:
            entries: Lista de entradas de historial.
            output_path: Ruta de salida .csv.

        Returns:
            True si fue exitoso.
        """
        try:
            with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
                if not entries:
                    f.write("Sin datos\n")
                    return True
                writer = csv.DictWriter(f, fieldnames=list(entries[0].to_dict().keys()))
                writer.writeheader()
                for entry in entries:
                    writer.writerow(entry.to_dict())
            logger.info("Reporte CSV exportado: %s", output_path)
            return True
        except OSError as exc:
            logger.error("Error al exportar CSV: %s", exc)
            return False

    def export_xlsx(self, entries: Sequence[HistoryEntry], output_path: Path) -> bool:
        """
        Exporta entradas a XLSX con formato profesional.

        Args:
            entries: Lista de entradas de historial.
            output_path: Ruta de salida .xlsx.

        Returns:
            True si fue exitoso.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte de Compresión"  # type: ignore[assignment]

            # Título
            ws.merge_cells("A1:J1")  # type: ignore[union-attr]
            title_cell = ws["A1"]  # type: ignore[index]
            title_cell.value = "FileOptimizer Pro — Reporte de Compresión"
            title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
            title_cell.fill = PatternFill("solid", fgColor="1E40AF")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 30  # type: ignore[union-attr]

            # Fecha de generación
            ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"  # type: ignore[index]
            ws["A2"].font = Font(italic=True, color="6B7280")  # type: ignore[index]

            if not entries:
                ws["A3"] = "Sin datos registrados."  # type: ignore[index]
                wb.save(output_path)
                return True

            # Encabezados
            headers = list(entries[0].to_dict().keys())
            header_row = 4
            header_fill = PatternFill("solid", fgColor="DBEAFE")
            thin_border = Border(
                bottom=Side(style="thin", color="93C5FD")
            )
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=header_row, column=col, value=header)
                cell.font = Font(bold=True, name="Calibri", size=10)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            # Datos
            for row_idx, entry in enumerate(entries, start=header_row + 1):
                row_data = list(entry.to_dict().values())
                fill_color = "F8FAFC" if row_idx % 2 == 0 else "FFFFFF"
                for col, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.fill = PatternFill("solid", fgColor=fill_color)
                    cell.font = Font(name="Calibri", size=9)
                    cell.alignment = Alignment(horizontal="center")

            # Ajustar anchos de columna
            for col in ws.columns:  # type: ignore[union-attr]
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)  # type: ignore[union-attr]

            wb.save(output_path)
            logger.info("Reporte XLSX exportado: %s", output_path)
            return True
        except Exception as exc:
            logger.error("Error al exportar XLSX: %s", exc)
            return False

    def export_pdf(self, entries: Sequence[HistoryEntry], output_path: Path) -> bool:
        """
        Exporta entradas a PDF con reportlab.

        Args:
            entries: Lista de entradas de historial.
            output_path: Ruta de salida .pdf.

        Returns:
            True si fue exitoso.
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            )

            doc = SimpleDocTemplate(
                str(output_path),
                pagesize=landscape(A4),
                leftMargin=1.5 * cm,
                rightMargin=1.5 * cm,
                topMargin=2 * cm,
                bottomMargin=2 * cm,
            )
            styles = getSampleStyleSheet()

            story = []

            # Título
            title_style = ParagraphStyle(
                "Title",
                parent=styles["Title"],
                fontSize=16,
                textColor=colors.HexColor("#1E40AF"),
                spaceAfter=6,
            )
            story.append(Paragraph("FileOptimizer Pro — Reporte de Compresión", title_style))

            subtitle_style = ParagraphStyle(
                "Sub",
                parent=styles["Normal"],
                fontSize=9,
                textColor=colors.HexColor("#6B7280"),
                spaceAfter=16,
            )
            story.append(Paragraph(
                f"Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}  |  "
                f"Total de registros: {len(entries)}",
                subtitle_style,
            ))

            if not entries:
                story.append(Paragraph("Sin datos registrados.", styles["Normal"]))
                doc.build(story)
                return True

            # Tabla
            headers = list(entries[0].to_dict().keys())
            table_data = [headers] + [list(e.to_dict().values()) for e in entries]

            col_widths = [3.5 * cm, 2 * cm, 2.5 * cm, 2.5 * cm, 2 * cm, 2.5 * cm, 2 * cm, 2 * cm, 3 * cm, 2 * cm]

            t = Table(table_data, colWidths=col_widths, repeatRows=1)
            t.setStyle(TableStyle([
                # Encabezado
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 7),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("TOPPADDING", (0, 0), (-1, 0), 6),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                # Datos
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 6.5),
                ("ALIGN", (0, 1), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EFF6FF")]),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#BFDBFE")),
                ("TOPPADDING", (0, 1), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
            ]))
            story.append(t)
            doc.build(story)
            logger.info("Reporte PDF exportado: %s", output_path)
            return True
        except Exception as exc:
            logger.error("Error al exportar PDF: %s", exc)
            return False
